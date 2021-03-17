import datetime

import xlrd
import xlwt
from xlutils.copy import copy
import json
from openpyxl import load_workbook


def change_dic(dic):
    '''
    将包含字典的集合转化成包含集合的集合
    '''
    buf = []
    if len(dic) == 0:
        return dic
    buf.append(list(dic[0].keys()))
    for each in dic:
        buf.append(list(each.values()))
    return buf


def change_dic_by_name(dic):
    dic = change_dic(dic)
    _dic = list()
    for _ in dic:
        if _[0] == "column_name":
            continue
        _dic.append(_[0])
    return [_dic]


# 保存常规excel文件：内容保存在第一页sheet中的第一列
def write_general_excel(datas, file_name):
    xls = xlwt.Workbook()
    sheet = xls.add_sheet('Tables_list')
    for row in range(len(datas)):
        sheet.write(row, 0, datas[row])
    xls.save(file_name)

def change_dif_data_by_key(dif_array, dif_row_excel, dif_column_excel, dif_row_data, dif_column_data, row_name):
    data1, data2 = [], []
    max_row = max(len(row_name[0]), len(row_name[1]))
    data1.append(['列'] + row_name[0] + [''] * (max_row - len(row_name[0])))
    data2.append(['列'] + row_name[1] + [''] * (max_row - len(row_name[1])))

    if dif_column_excel != [] or dif_column_data != []:
        try:
            len_excel = len(dif_column_excel)
        except IndexError:
            len_excel = 0
        try:
            len_data = len(dif_column_data)
        except IndexError:
            len_data = 0
        for col in range(max(len_excel, len_data)):
            buf_1, buf_2 = [col+1] + [''] * max_row, [col+1] + [''] * max_row
            data1.append(buf_1)
            data2.append(buf_2)

        for x, y, content in dif_column_data:
            data1[x] = data1[x][:y] + content

        for x, y, content in dif_column_excel:
            data2[x] = data2[x][:y] + content

        for x1, y1, x2, y2, one, two in dif_array:
            data1[x1][y1] = one
            data2[x1][y2] = two
            if data2[x1][0] != x2:
                data2[x1][0] = x2

        for x, content in dif_row_data:
            if x >= len(data1):
                data1.append([x] + content + [''] * (max_row - len(row_name[0])))
                data2.append([x] + [''] * max_row)
            else:
                data1[x][1:] = content
        for x, content in dif_row_excel:
            if x >= len(data2):
                data2.append([x] + content + [''] * (max_row - len(row_name[1])))
                data1.append([x] + [''] * max_row)
            else:
                data2[x][1:] = content
    elif dif_array != []:
        for col in range(dif_array[-1][0]):
            buf_1, buf_2 = [col + 1] + [''] * max_row, [col + 1] + [''] * max_row
            data1.append(buf_1)
            data2.append(buf_2)

        buf = list(range(1, 1 + dif_array[-1][0]))
        for x1, y1, x2, y2, one, two in dif_array:
            data1[x1][y1] = one
            data2[x1][y2] = two
            if x1 in buf:
                buf.remove(x1)
            if data2[x1][0] != x2:
                data2[x1][0] = x2
        for i in buf[::-1]:
            data1.pop(i)
            data2.pop(i)


        for x, content in dif_row_data:
            data1.append([x + 1] + content)
            data2.append([x + 1] + [''] * max_row)
        for x, content in dif_row_excel:
            data2.append([x + 1] + content)
            data1.append([x + 1] + [''] * max_row)
    else:
        for x, content in dif_row_data:
            if x >= len(data1):
                data1.append([x + 1] + content + [''] * (max_row - len(row_name[0])))
                data2.append([x + 1] + [''] * max_row)
        for x, content in dif_row_excel:
            if x >= len(data2):
                data2.append([x + 1] + content + [''] * (max_row - len(row_name[1])))
                data1.append([x + 1] + [''] * max_row)
    return data1, data2



def change_dif_data(dif_array, dif_row_excel, dif_column_excel, dif_row_data, dif_column_data, row_name):
    data1, data2 = [], []
    max_row = max(len(row_name[0]), len(row_name[1]))
    data1.append(['列'] + row_name[0] + [''] * (max_row - len(row_name[0])))
    data2.append(['列'] + row_name[1] + [''] * (max_row - len(row_name[1])))
    if dif_column_excel != [] or dif_column_data != []:
        try:
            len_excel = len(dif_column_excel[0][1])
        except IndexError:
            len_excel = 0
        try:
            len_data = len(dif_column_data[0][1])
        except IndexError:
            len_data = 0
        for col in range(1, max(len_excel, len_data)):
            buf_1, buf_2 = [col] + [''] * max_row, [col] + [''] * max_row
            for row, content in dif_column_data:
                buf_1[row + 1] = content[col]
            for row, content in dif_column_excel:
                buf_2[row + 1] = content[col]
            data1.append(buf_1)
            data2.append(buf_2)
        for x, y, one, two in dif_array:
            data1[x - 1][y] = one
            data2[x - 1][y] = two
        for x, content in dif_row_data:
            if x >= len(data1):
                data1.append([x] + content + [''] * (max_row - len(row_name[0])))
                data2.append([x] + [''] * max_row)
            else:
                data1[x][1:] = content
        for x, content in dif_row_excel:
            if x >= len(data2):
                data2.append([x] + content + [''] * (max_row - len(row_name[1])))
                data1.append([x] + [''] * max_row)
            else:
                data2[x][1:] = content
    elif dif_array != []:
        for col in range(dif_array[-1][0]):
            buf_1, buf_2 = [col + 1] + [''] * max_row, [col + 1] + [''] * max_row
            data1.append(buf_1)
            data2.append(buf_2)
        buf = list(range(1, 1 + dif_array[-1][0]))
        for x, y, one, two in dif_array:
            if x in buf:
                buf.remove(x)
            data1[x][y] = one
            data2[x][y] = two
        for i in buf[::-1]:
            data1.pop(i)
            data2.pop(i)
        for x, content in dif_row_data:
            data1.append([x + 1] + content)
            data2.append([x + 1] + [''] * max_row)
        for x, content in dif_row_excel:
            data2.append([x + 1] + content)
            data1.append([x + 1] + [''] * max_row)
    else:
        for x, content in dif_row_data:
            if x >= len(data1):
                data1.append([x + 1] + content + [''] * (max_row - len(row_name[0])))
                data2.append([x + 1] + [''] * max_row)
        for x, content in dif_row_excel:
            if x >= len(data2):
                data2.append([x + 1] + content + [''] * (max_row - len(row_name[1])))
                data1.append([x + 1] + [''] * max_row)
    return data1, data2


def write_excel(data_array, filename, name_key):
    '''
    data_array为一个字典，字典key为sheet名字，值为对应内容
    函数功能为将多个sheet页保存为excel格式文件名为filename
    name_key : sheet0中， 表名的rule
    '''
    if not data_array:
        return False
    xls = xlwt.Workbook()
    _index = 1
    # 先判断
    if "field" in data_array.keys():
        field_item = {"table_name": 0, "col_name": 1, "col_value": 2}
        sheet = xls.add_sheet('field')
        print(data_array['field'])
        field_index = 0
        for item in data_array['field']:
            for _key, _value in field_item.items():
                field_row = 1
                sheet.write(0, field_index + _value, _key)
                for _item in item[_key]:
                    sheet.write(field_row, field_index + _value, _item)
                    field_row += 1
            field_index += 3
        data_array.pop("field")
        _index = 2
    sheetNmae = "Tables"
    if isinstance(name_key[0], list):
        sheetNmae = "Tables-key"

    sheet = xls.add_sheet(sheetNmae)
    row = 0
    for name in data_array.keys():
        sheet.write(row, 0, row + _index)
        sheet.write(row, 1, name)
        if isinstance(name_key[0], list):
            assert name == name_key[row][0]
            sheet.write(row, 2, json.dumps(name_key[row][1]))
        row += 1
    key = _index
    for name in data_array.keys():
        try:
            if len(name) >= 31:
                sheet = xls.add_sheet(str(key))
            else:
                sheet = xls.add_sheet(name)
            key += 1
        except Exception as e:
            print(e)
            return False
        col = 0
        style = xlwt.XFStyle()
        style.num_format_str = 'yyyy/m/d h:mm:ss'
        for each_col in data_array[name]:
            row = 0
            for each_row in each_col:
                if isinstance(each_row, datetime.datetime):
                    sheet.write(col, row, each_row, style)
                elif isinstance(each_row, datetime.date):
                    sheet.write(col, row, str(each_row))
                else:
                    try:
                        sheet.write(col, row, each_row)
                    except Exception as err:
                        print(err)
                row += 1
            col += 1
    xls.save(filename)
    return True


# 获取某sheet某name列内容
def read_excel_columns(path, key, sheet_name=0):
    ex = xlrd.open_workbook(path)
    if isinstance(sheet_name, int):
        sheet = ex.sheet_by_index(0)
    else:
        sheet = ex.sheet_by_name(sheet_name)
    rows = sheet.row_values(0)
    if key not in rows:
        key = float(key)
        if key not in rows:
            return False
    index = rows.index(key)
    return sheet.col_values(index)


# 一次性读取母文件及所有表内容
def read_excel_mu_datas(path):
    names = dict()
    excelfile = xlrd.open_workbook(path, formatting_info=True)
    sheet = excelfile.sheet_by_index(0)
    if sheet.name == "field":
        names['field'] = 0
        sheet = excelfile.sheet_by_index(1)
    if sheet.ncols not in [2, 3]:
        return -2

    big_datas = dict()
    table_key = {'Tables-key': None}
    if sheet.ncols == 3:
        snap_key = dict()
        for row in range(sheet.nrows):
            names[sheet.cell_value(row, 1)] = int(sheet.cell_value(row, 0))
            snap_key[sheet.cell_value(row, 1)] = sheet.cell_value(row, 2)
        table_key["Tables-key"] = snap_key
    else:
        for row in range(sheet.nrows):
            names[sheet.cell_value(row, 1)] = int(sheet.cell_value(row, 0))

    for name in names.keys():
        if 'sheet' in name:
            return -2
    for name in names.keys():
        sheet = excelfile.sheet_by_index(names[name])
        buf = []
        for row in range(sheet.nrows):
            content = []
            for col in range(sheet.ncols):
                if sheet.cell_type(row, col) == 3:
                    content.append(xlrd.xldate_as_datetime(sheet.cell_value(row, col), 0))
                else:
                    content.append(sheet.cell_value(row, col))
            buf.append(content)
        big_datas[name] = buf
    big_datas.update(table_key)
    return big_datas


def read_excel_datas(path, index):
    '''
    excelfile = xlrd.open_workbook(path)
    sheet = excelfile.sheet_by_name(name)
    return sheet._cell_values
    '''
    excelfile = xlrd.open_workbook(path)
    if index == None:
        index = 0
    if isinstance(index, str):
        sheet = excelfile.sheet_by_name(index)
    else:
        sheet = excelfile.sheet_by_index(index)
    buf = []
    for row in range(sheet.nrows):
        content = []
        for col in range(sheet.ncols):
            if sheet.cell_type(row, col) == 3:
                content.append(xlrd.xldate_as_datetime(sheet.cell_value(row, col), 0))
            else:
                content.append(sheet.cell_value(row, col))
        buf.append(content)
    return buf


# 读取sheet names
def read_excel_sheets(path):
    return xlrd.open_workbook(path).sheet_names()


# 读取excel首行names
def read_cxcel_rows(path, name=None):
    ex = xlrd.open_workbook(path)
    if name == None:
        sheet = ex.sheet_by_index(0)
    else:
        sheet = ex.sheet_by_name(name)
    if sheet.ncols < 1:
        return -2
    return sheet.row_values(0)


# 读取母文件第一页sheet
def read_excel_names(path):
    excelfile = xlrd.open_workbook(path)
    sheet = excelfile.sheet_by_index(0)
    if sheet.ncols not in [2, 3]:
        sheet = excelfile.sheet_by_index(1)

        if sheet.ncols != 2:
            return -2
    names = dict()
    for row in range(sheet.nrows):
        names[sheet.cell_value(row, 1)] = int(sheet.cell_value(row, 0))
    for name in names.keys():
        if 'sheet' in name:
            return -2
    return names


# 读取字段下载方式文件（新足）
def read_field_table(path):
    result = []
    excelfile = xlrd.open_workbook(path)
    sheet = excelfile.sheet_by_index(0)
    if sheet.ncols % 3 != 0:
        return -1
    for i in range(int(sheet.ncols / 3)):
        interim = {}
        j = 0
        for _ in ["table_name", "col_name", "col_value"]:
            items = set(sheet.col_values(i * 3 + j)[1:])
            if '' in items:
                items.remove('')
            interim[_] = list(items)
            j += 1
        result.append(interim)
    return result


# 读取配置文件（table list）
def _read_save_names(path):
    excelfile = xlrd.open_workbook(path)
    sheet = excelfile.sheet_by_index(0)
    array = []

    if (sheet.nrows >= 500) or (sheet.nrows == 0) or (sheet.ncols >= 500):
        return -1
    for mcol in range(sheet.ncols):
        for content in sheet.col_values(mcol):
            if content != '' and content not in array:
                array.append(content)
    return array

# 读取配置文件（table list）
def read_save_names(path):
    excelfile = xlrd.open_workbook(path)
    sheet = excelfile.sheet_by_index(0)
    array = []
    if sheet.name == "List-key":
        if sheet.ncols != 2:
            return -2
        for row in range(sheet.nrows):
            array.append([sheet.cell_value(row, 0), int(sheet.cell_value(row, 1))])
        return array

    if (sheet.nrows >= 500) or (sheet.nrows == 0) or (sheet.ncols >= 500):
        return -1
    for mcol in range(sheet.ncols):
        for content in sheet.col_values(mcol):
            if content != '' and content not in array:
                array.append(content)
    return array

def modifyExcel(path, content):

    fileOpen = xlrd.open_workbook(path, formatting_info=True)
    data = copy(fileOpen)

    # 获取第一张表
    table = fileOpen.sheets()[0]

    tableRead = data.get_sheet(0)

    idx = 2

    if data.get_sheet(0).name == "Tables":
        data.get_sheet(0).name = "Tables-key"
    elif data.get_sheet(0).name == "List-key":
        idx = 1
    elif data.get_sheet(0).name != "Tables-key":
        idx = 1
        data.get_sheet(0).name = "List-key"

    for i in range(len(table._cell_values)):
        if table._cell_values[i][idx-1] == content[0]:
            tableRead.write(i, idx, json.dumps(content[1]))
            data.save(path)
            return [True, "success"]

    return [False, "保存失败, 未找到目标表: %s" % content[0]]




if __name__ == '__main__':
    '''
    buf = []
    for i in range(10):
        dic = dict()
        for j in ('one', 'two', 'three', 'four', 'five'):
            dic[j] = i
        buf.append(dic)
    buf_1 = dict()
    buf_2 = dict()
    buf_1['list1'] = change_dic(buf)
    buf_1['list2'] = change_dic(buf)
    write_excel(buf_1)
    '''
    # array = read_save_names('C:/Users/admin/AppData/Local/Programs/Python/Python37/python_py/test.xls')
    # for each in array:
    #     print(each)
    # content = read_excel_columns(r"C:\Users\Administrator\Desktop\data.xlsx", 1, sheet_name="字段表")
    modifyExcel(r'C:\Users\Administrator\Desktop\test_1.xlsx', ['test', {"1": 1, "2": 8888888, "3": 3}])
    pass
